from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl
from constants import globalConstants as cons
from test_Sauce import Test_SauceDemo

class Test_Sauce_Optioanl:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(cons.BASE_URL)
        self.driver.maximize_window()
    
    def teardown_method(self):
        self.driver.quit()

    def getInvalidUserTestData():
        excelFile = openpyxl.load_workbook("data/login.xlsx")
        sheet= excelFile["invalid_login"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data

    @pytest.mark.parametrize("username,password", Test_SauceDemo.getValidUserTestData())
    def test_checkToCard(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.PASSWORD_ID)))
        passwordInput.send_keys(password)
        passwordInput.send_keys(Keys.ENTER)
        clickItem = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, cons.PRODUCT_XPATH)))
        clickItem.click()
        clickCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,cons.SHOPPING_CART_XPATH)))
        clickCart.click()
        checkItem = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, cons.CART_ITEM_XPATH)))
        assert checkItem.text == cons.ITEM_TITLE
    @pytest.mark.xfail
    @pytest.mark.parametrize("username,password",Test_SauceDemo.getValidUserTestData())
    def test_checkToCardRemoved(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.PASSWORD_ID)))
        passwordInput.send_keys(password)
        passwordInput.send_keys(Keys.ENTER)
        addItem = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,cons.PRODUCT_XPATH)))
        addItem.click()
        clickCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,cons.SHOPPING_CART_XPATH)))
        clickCart.click()
        removeItem =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"remove-sauce-labs-backpack")))
        removeItem.click()
        continueShopping =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"continue-shopping")))
        continueShopping.click()
        clickCart.click()
        checkItem = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, cons.CART_ITEM_XPATH)))
        assert checkItem.text == cons.ITEM_TITLE

    @pytest.mark.parametrize("username,password", getInvalidUserTestData())
    def test_failedLogin(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.PASSWORD_ID)))
        passwordInput.send_keys(password)
        passwordInput.send_keys(Keys.ENTER)
        errorMessage = self.driver.find_element(By.XPATH, cons.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
