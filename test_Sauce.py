from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from constants import globalConstants as cons
import openpyxl
import pytest
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

class Test_SauceDemo:
    def setup_method(self): 
    
        self.driver = webdriver.Chrome()
        self.driver.get(cons.BASE_URL)
        self.driver.maximize_window()

        
    def teardown_method(self):
        self.driver.quit()

    def getInvalidUserTestData():
        excelFile = openpyxl.load_workbook("data/login.xlsx")
        sheet= excelFile["locked_out"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data
  
    def getOnlyUsernameTestData():
        excelFile = openpyxl.load_workbook("data/login.xlsx")
        sheet = excelFile["only_username"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = sheet.cell(i,1).value
            data.append(username)
        return data
  
    def getValidUserTestData():
        excelFile = openpyxl.load_workbook("data/login.xlsx")
        sheet = excelFile["valid_login"]
        rows = sheet.max_row
        data = []
        for i in range(2, rows + 1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data

        
    @pytest.mark.parametrize("username, password", getInvalidUserTestData())
    def test_invalid_login(self, username, password):

        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,cons.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, cons.LOGIN_BTN_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,cons.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
    
    def test_username_and_password_required(self):

        loginBtn = self.driver.find_element(By.ID, cons.LOGIN_BTN_ID)
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, cons.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == "Epic sadface: Username is required"
    
    @pytest.mark.parametrize("username",getOnlyUsernameTestData())
    def test_password_required(self,username):

        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, cons.USERNAME_ID)))
        usernameInput.send_keys(username)

        loginBtn = self.driver.find_element(By.ID, cons.LOGIN_BTN_ID)
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH, cons.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == "Epic sadface: Password is required"

    @pytest.mark.parametrize("username,password",getValidUserTestData())
    def test_succesfully_login(self,username,password):

        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, cons.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, cons.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, cons.LOGIN_BTN_ID)
        loginButton.click()

        items = self.driver.find_elements(By.CLASS_NAME, cons.ITEM_CLASS_NAME)
        assert len(items)==6

